import pyrebase
import json
import base64
import pandas as pd
import os
from tabulate import tabulate
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad

# Firebase Configuration (Sensitive credentials redacted)
firebaseConfig = {
    "apiKey": "*******",
    "authDomain": "*******",
    "databaseURL": "*******",
    "projectId": "*******",
    "storageBucket": "*******",
    "messagingSenderId": "*******",
    "appId": "*******"
}

# Initialize Firebase
firebase = pyrebase.initialize_app(firebaseConfig)
db = firebase.database()

# AES Configuration (Sensitive key/IV redacted)
AES_KEY = b"*******"  # 16-byte key
AES_IV = b"*******"   # 16-byte IV

def decrypt_data(encrypted_b64, class_name, date, identifier):
    try:
        encrypted_data = base64.b64decode(encrypted_b64)
        cipher = AES.new(AES_KEY, AES.MODE_CBC, AES_IV)
        padded_data = cipher.decrypt(encrypted_data)
        plaintext = unpad(padded_data, AES.block_size)
        json_data = json.loads(plaintext.decode('utf-8'))
        print(f"Decrypted data for {class_name}/{date}/{identifier}: {json_data}")
        return json_data
    except Exception as e:
        print(f"Decryption error for {class_name}/{date}/{identifier}: {e}")
        return None

def fetch_and_process_attendance():
    attendance_records = db.child("attendance_records").get().val()
    all_csv_data = []
    for class_name in attendance_records:
        class_data = attendance_records[class_name]
        for date in class_data:
            date_data = class_data[date]
            csv_data = []
            if isinstance(date_data, dict):
                for id_key, record in date_data.items():
                    if id_key.startswith("id_"):
                        id_num = id_key.replace("id_", "")
                    else:
                        id_num = id_key
                    if isinstance(record, str):
                        encrypted_b64 = record
                        decrypted = decrypt_data(encrypted_b64, class_name, date, id_key)
                        if decrypted:
                            csv_data.append({
                                "id": decrypted["id"],
                                "matric_number": decrypted["matric_number"],
                                "course": decrypted["course"],
                                "date": decrypted["date"],
                                "present": decrypted["present"],
                                "present_time": decrypted["present_time"],
                                "absent_time": decrypted["absent_time"] if decrypted["absent_time"] else ""
                            })
            elif isinstance(date_data, list):
                for idx, record in enumerate(date_data):
                    if record:
                        encrypted_b64 = record
                        decrypted = decrypt_data(encrypted_b64, class_name, date, idx)
                        if decrypted:
                            csv_data.append({
                                "id": decrypted["id"],
                                "matric_number": decrypted["matric_number"],
                                "course": decrypted["course"],
                                "date": decrypted["date"],
                                "present": decrypted["present"],
                                "present_time": decrypted["present_time"],
                                "absent_time": decrypted["absent_time"] if decrypted["absent_time"] else ""
                            })
            if csv_data:
                df = pd.DataFrame(csv_data)
                output_dir = f"attendance/{class_name}"
                os.makedirs(output_dir, exist_ok=True)
                csv_path = f"{output_dir}/{date}.csv"
                df.to_csv(csv_path, index=False)
                print(f"Generated CSV: {csv_path}")
                all_csv_data.extend(csv_data)

    if all_csv_data:
        all_df = pd.DataFrame(all_csv_data)
        all_df.sort_values(by=["course", "date", "id"], inplace=True)
        print("\nAll Attendance Data:")
        print(tabulate(all_df, headers="keys", tablefmt="psql", showindex=False))

        # Generate HTML Report
        html_path = "attendance/attendance_report.html"
        html_content = """
        <html>
        <head>
            <title>Attendance Report</title>
            <style>
                table { border-collapse: collapse; width: 100%; }
                th, td { border: 1px solid black; padding: 8px; text-align: left; }
                th { background-color: #f2f2f2; }
            </style>
        </head>
        <body>
            <h2>Attendance Report</h2>
            <table>
                <tr>
                    <th>ID</th><th>Matric Number</th><th>Course</th><th>Date</th>
                    <th>Present</th><th>Present Time</th><th>Absent Time</th>
                </tr>
        """
        for _, row in all_df.iterrows():
            html_content += f"""
                <tr>
                    <td>{row['id']}</td><td>{row['matric_number']}</td><td>{row['course']}</td>
                    <td>{row['date']}</td><td>{row['present']}</td><td>{row['present_time']}</td>
                    <td>{row['absent_time']}</td>
                </tr>
            """
        html_content += "</table></body></html>"
        with open(html_path, "w") as f:
            f.write(html_content)
        print(f"Generated HTML: {html_path}")

        # Generate Excel Report
        excel_path = "attendance/attendance_summary.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Summary"
        headers = ["ID", "Matric Number", "Course", "Date", "Present", "Present Time", "Absent Time"]
        ws.append(headers)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.border = thin_border
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for _, row in all_df.iterrows():
            ws.append([row["id"], row["matric_number"], row["course"], row["date"],
                       row["present"], row["present_time"], row["absent_time"]])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left")
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
        wb.save(excel_path)
        print(f"Generated Excel: {excel_path}")

if __name__ == "__main__":
    fetch_and_process_attendance()
